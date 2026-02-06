from pathlib import Path
from PySide6.QtWidgets import QTableWidget
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from PySide6.QtCore import Qt

fontHdr = Font(name='Verdana', size=16, bold=True)
fontCell = Font(name='Verdana', size=14)

def qtablewidget_to_xlsx(
    table: QTableWidget,
    filepath: str | Path,
    sheet_name: str = "Sheet1"
) -> None:
    filepath = str(filepath)

    wb = Workbook()
    ws = wb.active
    # Workbook window hints
    view = wb.views[0]
    view.windowWidth = 18000
    view.windowHeight = 10000
    view.xWindow = 200
    view.yWindow = 200
    view.minimized = False

    # Sheet presentation
    ws.sheet_view.zoomScale = 85
    ws.sheet_view.showGridLines = False

    ws.title = sheet_name

    rows = table.rowCount()
    cols = table.columnCount()

    # ---- headers ----
    for c in range(cols):
        header_item = table.horizontalHeaderItem(c)
        ws.cell(row=1, column=c + 1).value = (
            header_item.text() if header_item else ""
        )
        ws.cell(row=1, column=c +1).font = fontHdr
    # row = ws.row_dimensions[rows]
    # col = ws.column_dimensions[cols]
    # for r in range(1, rows + 1):
    #     for c in range(1, cols + 1):
    #         ws[r,c].alignment = Alignment(horizontal="center", vertical="center")
    #         ws[r,c].font = Font(bold=True)


    # ---- data ----
    for r in range(rows):
        for c in range(cols):
            item = table.item(r, c)
            cell = ws.cell(row=r +2, column=c +1)
            cell.value = cell_value(item)
            cell.font = fontCell
            # ws.cell(row=r + 2, column=c + 1).value = cell_value(item)

    # ---- autosize columns (approx) ----
    autosize_columns(ws)

    # for c in range(1, cols + 1):
    #     max_len = 0
    #     for r in range(1, rows + 2):
    #         v = ws.cell(row=r, column=c).value
    #         if v is not None:
    #             max_len = max(max_len, len(str(v)))
    #     ws.column_dimensions[get_column_letter(c)].width = min(max_len + 2, 60)

    # Optional Excel niceties
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)

# def autosize_columns(ws, min_width=10, max_width=50):
#     for col_cells in ws.columns:
#         max_length = 0
#         col_letter = get_column_letter(col_cells[0].column)
#
#         for cell in col_cells:
#             if cell.value is not None:
#                 max_length = max(max_length, len(str(cell.value)))
#
#         adjusted = max(min_width, min(max_length + 2, max_width))
#         ws.column_dimensions[col_letter].width = adjusted

def autosize_columns(ws, font_size=16):
    scale = font_size / 11

    for col_cells in ws.columns:
        max_length = max(
            (len(str(cell.value)) for cell in col_cells if cell.value),
            default=0
        )

        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(8, (max_length + 2) * scale)

def cell_value(item):
    if item is None:
        return ""
    # Prefer raw value if you stored one
    raw = item.data(Qt.UserRole)
    return raw if raw is not None else item.text()